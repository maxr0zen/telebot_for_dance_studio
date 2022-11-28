import telebot
from openpyxl import load_workbook
from telebot import types

bot = telebot.TeleBot('5743624493:AAH1KRDwhERIsIeDCJ5U0jDyLQS5k6urMkk')

class OpenGroups:
    info = type(str)
    button = type(telebot.types.KeyboardButton)
    text = type(str)
    counter = 0
    name = type(str)
    all_keys = {}

    def __init__(self, info, button, text,name):
        OpenGroups.counter += 1
        self.info = info
        self.button = button
        self.text = text
        self.name = name
        self.text_overflow(text, info= self.info, all_keys= OpenGroups.all_keys)

    def text_overflow(self, text, info, all_keys):
        self.all_keys[text] = info


group_4_5 = OpenGroups(info = 'Основной педагог: Корсакова Евгения \n Расписание: \n Понедельник: 18:00 - 19:00 \n Четверг: 18:30 - 19:30 \n Cуббота: 16:00 - 17:00',
                       button = types.KeyboardButton("/📅 группа 4-5"), text = "/📅 группа 4-5", name= 'group_4_5')
group_6_7 = OpenGroups(info = 'Основной педагог: Иващенко Диана \n Расписание: \n Понедельник:  19:00 – 20:00 \n Вторник: 19:00 – 20:00\n Четверг:  19:30 – 20:30\n Суббота:  15:00 – 16:00',
                       button = types.KeyboardButton("/📅 группа 6-7"), text = "/📅 группа 6-7", name= 'group_6_7')
group_8_9 = OpenGroups(info = 'Основной педагог: Агафонов Александр\n Расписание:\n Понедельник:  18:00 – 19:00\n Среда: 18:15 – 19:15\n Пятница:  18:00 – 19:00\n Суббота:  11:00 – 12:00' + '\n\n' +
                       'Основной педагог: Агафонов Александр\n Расписание:\n Понедельник:  19:00 – 20:00\n Вторник: 18:15 – 19:30 \n Среда:  19:15 – 20:30\n Пятница:  19:00 – 20:15',
                       button = types.KeyboardButton("/📅 группа 8-9"), text = "/📅 группа 8-9", name= 'group_8_9')
group_9_10 = OpenGroups(info = 'Основной педагог: Коваленко Александр\n Расписание:\n Понедельник:  16:00 – 17:30\n Среда: 16:30 – 18:00\n Четверг:  17:00 – 18:30\n Суббота : 12:00 – 13:00; 13:00 – 14:30',
                        button =  types.KeyboardButton("/📅 группа 9-10"), text = "/📅 группа 9-10", name= 'group_9_10')
group_13_16 = OpenGroups(info = 'Основной педагог: Задворнева Виктория\n Расписание:\n Вторник:  16:00 – 17:30\n Среда: 16:00 – 17:30\n Пятница:  16:00 – 17:30\n Суббота:  17:00 – 18:30',
                         button = types.KeyboardButton("/📅 группа 13-16"), text = "/📅 группа 13-16", name= 'group_13_16')
group_Moms = OpenGroups(info = 'Основной педагог: Скобелкина Лариса\n Расписание: \n Вторник: 19:30 – 21:00\n Суббота:  13:00 – 14:30',
                        button = types.KeyboardButton("/📅 группа для Мам"), text = "/📅 группа для Мам", name='group_Moms')
group_Tvorcha = OpenGroups(info = 'Основной педагог: Мария Веселова\n Расписание: \n Вторник:  18:00 – 19:30 \n Четверг:  18:00 – 19:30\n Суббота:  18:00 – 19:30',
                           button = types.KeyboardButton("/📅 творческая 12+"), text = "/📅 творческая 12+", name='group_Tvorcha')
group_Utro = OpenGroups(info = 'Основной педагог: Коваленко Александр\n Расписание:\n Вторник:  10:00 – 11:30\n Среда: 10:00 – 11:30\n Пятница: 10:00 – 11:30\n Суббота:  12:00 – 13:00',
                        button = types.KeyboardButton("/📅 утренняя 10-14"), text = "/📅 утренняя 10-14", name='group_Utro')



wb = load_workbook('./Gruppy_Vzroslye.xlsx')                            #Сичтвываем первый excel файл
some_lists = wb.get_sheet_names()                                       #Считываем листы с excel
phones = []                                                             #Массив телефонов
till_classes = []                                                       #Массив оставшихся занятий
names = []                                                              #Массив Фамилии и Имяни ученика студии
alert = ''                                                              #Переменная для формирования строки для data.txt
parsed_data = []                                                        #Переменная для считывания строки из data.txt
table_button = '/📅 Расписание'
chose_group_button = '👀Посмотреть сколько занятий осталось'
group_name = {}
sheet_titles = []
childs_count = []
all_data = []

def list_manager(some_lists,phones,till_classes,names,wb,group_name,sheet_titles,childs_count, all_data):              #Функция, которая парсит excel и записывает в
    print('---Reading excel---')                                                   #соответствующие массивы данные
    current_group = {}
    for item in range(len(some_lists)):
        sheet = wb.get_sheet_by_name(some_lists[item])
        print(sheet.title)
        sheet_titles.append(sheet.title)
        number_of_childs = sheet['Y4'].value
        childs_count.append(number_of_childs)
        current_data = []
        for rower in range(5, number_of_childs + 5):
            phones.append(str(sheet.cell(row=rower, column=23).value))
            till_classes.append(str(sheet.cell(row=rower, column=24).value))
            names.append(sheet.cell(row=rower, column=2).value)
            current_name = sheet.cell(row=rower, column=2).value
            current_TK = str(sheet.cell(row=rower, column=24).value)
            current_data += [current_name, current_TK]
        current_group = {sheet.title: current_data}
        group_name.update(current_group)
        print(current_group)
    return print('---Done---'), phones, till_classes, names, group_name, sheet_titles,current_group



def little_lessons_count(phones):                   #Проверка, если у человека меньше 3 занятий, то
    data2 = open('./data.txt', 'r+')                #приходит уведомление о продлении абонемента
    parsed_data = data2.readlines()
    for item in range(len(parsed_data)):
        now_str = parsed_data[item].split()
        now_phone = now_str[0]
        print(now_phone)
        if (till_classes[item] <= 3) and (int(now_phone) in phones):
            bot.send_message(now_str[1], 'У вас осталось меньше 3 занятий, поажлуйста, продлите абонемент')
    data2.close()
    little_lessons_count(phones)


list_manager(some_lists,phones, till_classes, names,wb, group_name, sheet_titles, childs_count,all_data)                 #Запуск функции от второго файла
wb = load_workbook('./Gruppy_Deti.xlsx')
some_lists = wb.get_sheet_names()
print(some_lists)
list_manager(some_lists, phones, till_classes, names, wb, group_name, sheet_titles, childs_count,all_data)

current_id = 0
print(group_name)




class AllGroup:
    text = type(str)
    button = type(str)
    members = type(list)
    till_classes = type(list)
    info = type(list)
    def __init__(self, text, button, info, members):
        self.text = text
        self.button = button
        self.info = info
        self.members_collect(members= members, info= info, till_classes= info)
    def members_collect(self, members, info, till_classes):
        members :list = []
        till_classes :list = []
        for value in range(len(info)):
            if value%2 == 0:  members.append(info[value])
            else: till_classes.append(info[value])
        print(members)
        print(till_classes)




for key, value in group_name.items():
    print(key)
    group = AllGroup(text= key, button= types.KeyboardButton(value), info= value, members= None)

@bot.message_handler(commands=['start'])                                #Стартовое сообщение бота 
def greetings(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn1 = types.KeyboardButton(table_button)
    btn2 = types.KeyboardButton(chose_group_button)
    markup.add(btn1, btn2)
    bot.send_message(message.from_user.id, "бот запущен", reply_markup=markup)



@bot.message_handler(content_types=['text'])                            #Основной функционал бота
def get_text_messages(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn1 = types.KeyboardButton(table_button)
    btn2 = types.KeyboardButton(chose_group_button)
    markup.add(btn1, btn2)
    to_start = types.KeyboardButton("К началу")


    current_name = str(message.text)
    open_group_previous = types.KeyboardButton(".◀")
    open_group_next = types.KeyboardButton(".▶")

    open_group_table_first_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    open_group_table_first_markup.add(group_4_5.button, group_6_7.button, group_8_9.button, group_9_10.button, open_group_next)

    open_group_table_second_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    open_group_table_second_markup.add(group_13_16.button,group_Moms.button, group_Tvorcha.button, group_Utro.button, open_group_previous, to_start)
    all_groups_first_table = types.ReplyKeyboardMarkup(resize_keyboard=True)
    all_groups_2_table = types.ReplyKeyboardMarkup(resize_keyboard=True)
    all_groups_3_table = types.ReplyKeyboardMarkup(resize_keyboard=True)
    all_groups_4_table = types.ReplyKeyboardMarkup(resize_keyboard=True)
    Vika_closed = types.KeyboardButton("Закрытая Вика")
    gr_13_16 = types.KeyboardButton("13-16")
    X_bpm = types.KeyboardButton("X-bpm")
    P_teens = types.KeyboardButton("P teens")
    K_lab = types.KeyboardButton("K-lab")
    The_way = types.KeyboardButton("The-way")
    Arkies = types.KeyboardButton("Arkies")
    Tvorcha = types.KeyboardButton('Творча')
    baepsae = types.KeyboardButton("Baepsae")
    moms = types.KeyboardButton("Мамы")
    utro = types.KeyboardButton("Утро")
    gr_4_5 = types.KeyboardButton("Группа 4-5")
    gr_6_7 = types.KeyboardButton("Группа 6-7")
    gr_8_9_1 = types.KeyboardButton("Группа1 8-9")
    gr_8_9_2 = types.KeyboardButton("Группа2 8-9")
    from1to2 = types.KeyboardButton("Вторая ▶")
    from2to3 = types.KeyboardButton("Третья ▶")
    from3to4 = types.KeyboardButton("Четвертая ▶")
    from4to3 = types.KeyboardButton("◀ Третья")
    from3to2 = types.KeyboardButton("◀ Вторая")
    from2to1 = types.KeyboardButton("◀ Первая")
    all_groups_first_table.add(Vika_closed, gr_13_16, X_bpm, P_teens, from1to2)
    all_groups_2_table.add(K_lab,The_way,Arkies,from2to1,Tvorcha,from2to3)
    all_groups_3_table.add(baepsae,moms,utro,from3to2,gr_4_5,from3to4)
    all_groups_4_table.add(gr_6_7,gr_8_9_1,gr_8_9_2,from4to3)

    if current_name in sheet_titles:
        bot.send_message(message.from_user.id, "Введите Имя и Фамилию")
        now = group_name[current_name]
        print(now)
    if message.text == table_button or message.text == ".◀":
        bot.send_message(message.from_user.id, "Выберите группу", reply_markup=open_group_table_first_markup)
    if message.text == ".▶":
        bot.send_message(message.from_user.id, "Другие Группы", reply_markup=open_group_table_second_markup)
    elif message.text in OpenGroups.all_keys:
        for item in OpenGroups.all_keys:
            if message.text == item:
                bot.send_message(message.from_user.id, OpenGroups.all_keys[item])

    if message.text == "👀Посмотреть сколько занятий осталось" or message.text == "◀ Первая":
        bot.send_message(message.from_user.id,"Первая страница",reply_markup=all_groups_first_table)
    if message.text == "К началу":
        bot.send_message(message.from_user.id, "Возвращаюсь к началу", reply_markup=markup)
    if message.text == "Вторая ▶" or message.text == "◀ Вторая":
        bot.send_message(message.from_user.id, "Вторая страница",reply_markup= all_groups_2_table)
    if message.text == "Третья ▶" or message.text == "◀ Третья":
        bot.send_message(message.from_user.id, "Третья страница",reply_markup= all_groups_3_table)
    if message.text == "Четвертая ▶":
        bot.send_message(message.from_user.id, "Четвертая страница",reply_markup= all_groups_4_table)
    if current_name in names:
        for i in range(len(names)):
            if current_name == names[i]:
                bot.send_message(message.from_user.id, F"У вас осталось {till_classes[i]} занятий")









if __name__ == '__main__':
    bot.polling(none_stop=True, interval=0)




















